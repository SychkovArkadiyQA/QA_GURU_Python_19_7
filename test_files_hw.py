import csv
import pytest
import os.path
import zipfile
from pypdf import PdfReader
from openpyxl import load_workbook
from io import TextIOWrapper


CURRENT_FILE = os.path.abspath(__file__) # получаем абсолютный путь к текущему файлу
CURRENT_DIRECTORY = os.path.dirname(CURRENT_FILE) # получаем абсолютный путь к текущей директории
TMP_DIR = os.path.join(CURRENT_DIRECTORY, 'files') # делаем склейку пути к текущей директории и папке files
ZIP_DIR = os.path.join(CURRENT_DIRECTORY, 'resources') # делаем склейку пути к текущей директории и папке resources

# TODO Реализовать чтение и проверку содержимого каждого файла из архива не распаковывая сам архив;

def test_csv(create_zip):
    archive_path = os.path.join(ZIP_DIR, 'archive.zip')
    with zipfile.ZipFile(archive_path) as zip_file:
        with zip_file.open('CSV_for_test.csv') as csv_file:
            csvreader = list(csv.reader(TextIOWrapper(csv_file, 'utf-8-sig')))
            print(csvreader)
            assert csvreader == [['Anna', 'Pavel', 'Peter'],
                                 ['Alex', 'Serj', 'Yana']]

def test_xlsx(create_zip):
    archive_path = os.path.join(ZIP_DIR, 'archive.zip')
    with zipfile.ZipFile(archive_path) as zip_file:
        with zip_file.open('XLSX_for_test.xlsx') as xlsx_file:
            excel_value = load_workbook(xlsx_file).active  # получаем активный лист
            # Проверяем значения конкретных ячеек
            assert excel_value.cell(row=3, column=2).value == 'Mara'
            assert excel_value.cell(row=3, column=3).value == 'Hashimoto'
            assert excel_value.cell(row=3, column=8).value == 1582

def test_pdf(create_zip):
    archive_path = os.path.join(ZIP_DIR, 'archive.zip')
    with zipfile.ZipFile(archive_path) as zip_file:
        with zip_file.open('PDF_for_test.pdf') as pdf_file:
            pdfreader = PdfReader(pdf_file)
            text = pdfreader.pages[0].extract_text()
            print(text)
            assert text == 'PDF for test'

